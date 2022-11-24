#!/usr/bin/env python3

# OpenBACH is a generic testbed able to control/configure multiple
# network/physical entities (under test) and collect data from them. It is
# composed of an Auditorium (HMIs), a Controller, a Collector and multiple
# Agents (one for each network entity that wants to be tested).
#
#
# Copyright © 2016-2020 CNES
#
#
# This file is part of the OpenBACH testbed.
#
#
# OpenBACH is a free software : you can redistribute it and/or modify it under
# the terms of the GNU General Public License as published by the Free Software
# Foundation, either version 3 of the License, or (at your option) any later
# version.
#
# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY, without even the implied warranty of MERCHANTABILITY or FITNESS
# FOR A PARTICULAR PURPOSE. See the GNU General Public License for more
# details.
#
# You should have received a copy of the GNU General Public License along with
# this program. If not, see http://www.gnu.org/licenses/.


"""Provide time period summary of data generated by OpenBACH jobs"""


__author__ = 'Viveris Technologies'
__credits__ = '''Contributors:
 * Aichatou Garba Abdou <aichatou.garba-abdou@viveris.fr>
'''


import os
import syslog
import argparse
import tempfile
import itertools
from datetime import datetime

import pandas as pd
from dateutil.parser import parse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Side, Border, Font, PatternFill

import collect_agent
from data_access.post_processing import Statistics


UNIT_OPTION={'s', 'ms' ,'bits/s', 'Kbits/s', 'Mbits/s','Gbits/s','Bytes' ,'KBytes', 'MBytes', 'GBytes'}


def worksheet_style(worksheet, title):
    worksheet.title = title
    fill_color = PatternFill(fill_type='solid',start_color='00333399',end_color='FF000000')
    border_color = Side(border_style='thin', color='000000')

    for index, column in enumerate(worksheet.columns, 1):
        column_letter = get_column_letter(index)
        worksheet['{}1'.format(column_letter)].fill = fill_color
        worksheet['{}1'.format(column_letter)].font = Font(size=12, bold=True, color='00FFFFFF')
        for cell in column:
            cell.border = Border(top=border_color, bottom=border_color, left=border_color, right=border_color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        worksheet.column_dimensions[column_letter].width = 20

    for index, _ in enumerate(worksheet.rows, 1):
        A = get_column_letter(1)
        worksheet.row_dimensions[index].height = 30
        worksheet['{}{}'.format(A, index)].fill = fill_color
        worksheet['{}{}'.format(A, index)].font = Font(size=12, bold=True, color='00FFFFFF')


def _get_column_letter(worksheet, value):
    for index, column in enumerate(worksheet.columns, 1):
        for cell in column:
            if cell.value == value:
                return get_column_letter(index)
            # Only check first cell (header) of each column
            break


def get_evol_value(filepath, field, stats):
    workbook = load_workbook(filepath)
    worksheet = workbook[field]
    column = _get_column_letter(worksheet, stats)
    if column is not None:
        it = iter(worksheet[column])
        next(it)  # Skip header
        return [cell.value for cell in it]
    return []


def reference_style(worksheet, percentage, column_title):
    column = _get_column_letter(worksheet, column_title)
    if column is None:
        return

    row = len(worksheet[column])  # Get last row index
    if percentage <= 33:
        font_color = "ff0000"
    elif percentage <= 66:
        font_color = "ff7f00"
    else:
        font_color = "00ff00"

    worksheet[f'{column}{row}'].font = Font(color=font_color)


def get_trend(stability_threshold, value, reference):
    threshold = reference * stability_threshold / 100
    difference = value - reference

    if difference < -threshold:
        state = '\u2198'  # Down arrow
    elif difference <= threshold:
        state = '\u268C'  # Equal sign
    else:
        state = '\u2197'  # Up arrow

    return difference * 100 / reference, state


def multiplier(base, unit):
        if unit == base:
                return 1
        if unit.startswith('GBytes'):
                return 1024 * 1024 * 1024
        if unit.startswith('MBytes'):
                return 1024 * 1024
        if unit.startswith('KBytes'):
                return 1024
        if unit.startswith('m'):
                return 0.001
        if unit.startswith('s'):
                return 1000
        if unit.startswith('Gbits'):
                return 1000 * 1000 * 1000
        if unit.startswith('Mbits'):
                return 1000 * 1000
        if unit.startswith('Kbits'):
                return 1000

        return 1


def main(
        agent_name, job_name, statistic_name, timestamp_boundaries,
        start_day, start_evening, start_night,
        reference, stability_threshold, stat_unit, table_unit,
        path_to_file, stat_title, compute_median, compute_mean, stats_with_suffixes):

    statistics = Statistics.from_default_collector()
    statistics.origin = 0
    with tempfile.TemporaryDirectory(prefix='openbach-summary_time_period-') as root_folder:
        if not timestamp_boundaries:
            timestamp = None
        else:
            begin_date, end_date = map(parse, timestamp_boundaries)
            timestamp = [int(begin_date.timestamp() * 1000), int(end_date.timestamp() * 1000)]

        data_collection = statistics.fetch_all(
                job=job_name, agent=agent_name,
                suffix = None if stats_with_suffixes else '',
                fields=[statistic_name],timestamps=timestamp)

        workbook = Workbook()
        worksheet = workbook.active

        if stability_threshold is None:
            stability_threshold = 10
        if start_day is None:
            start_day = 7
        if start_evening is None:
            start_evening = 18
        if start_night is None:
            start_night = 0

        scale_factor = 1 if stat_unit is None else multiplier(stat_unit, table_unit or stat_unit)

        means = data_collection.compute_function('mean', scale_factor, start_day, start_evening, start_night).round(2)
        medians = data_collection.compute_function('median', scale_factor, start_day, start_evening, start_night).round(2)

        means_ref = get_evol_value(path_to_file, statistic_name, 'Moyenne') if path_to_file else []
        medians_ref = get_evol_value(path_to_file, statistic_name, 'Médiane') if path_to_file else []

        header = [f'{statistic_name} ({table_unit})' if table_unit else statistic_name]
        if compute_mean:
            header.extend(('Moyenne', '% Moyenne Cible'))
            if path_to_file:
                header.append('Évolution de la Moyenne')
        if compute_median:
            header.extend(('Mediane', '% Médiane Cible'))
            if path_to_file:
                header.append('Évolution de la Médiane')

        worksheet.append(header)

        for moment, mean, mean_ref, median, median_ref in itertools.zip_longest(means.index, means, means_ref, medians, medians_ref):
            row = [moment]
            if compute_mean:
                mean_percent = mean * 100 / reference
                row.extend((mean, mean_percent))
                if path_to_file:
                    if mean_ref:
                        trend, state = get_trend(stability_threshold, mean, mean_ref)
                        mean_trend = f'{state} {evol}%'
                    else:
                        mean_trend = 'NaN'
                    row.append(mean_trend)
            if compute_median:
                median_percent = median * 100 / reference
                row.extend((median, median_percent))
                if path_to_file:
                    if median_ref:
                        trend, state = get_trend(stability_threshold, median, median_ref)
                        median_trend = f'{state} {evol}%'
                    else:
                        median_trend = 'NaN'
                    row.append(median_trend)

            worksheet.append(row)

            if compute_mean:
                reference_style(worksheet, mean_percent, '% Moyenne Cible')
            if compute_median:
                reference_style(worksheet, median_percent, '% Médiane Cible')

        worksheet_style(worksheet, stat_title or statistic_name)
        filepath = os.path.join(root_folder, 'summary_time_period_{}.xlsx'.format(statistic_name))
        workbook.save(filepath)
        collect_agent.store_files(collect_agent.now(), figure=filepath)


if __name__ == '__main__':
    with collect_agent.use_configuration('/opt/openbach/agent/jobs/summary_time_period/summary_time_period_rstats_filter.conf'):
        parser = argparse.ArgumentParser(description=__doc__)
        parser.add_argument(
                'agent', metavar='AGENT_NAME',
                help='Agent name to fetch data from')
        parser.add_argument(
                'job', metavar='JOB_NAME',
                help='Job name to fetch data from')
        parser.add_argument(
                'statistic', metavar='STATISTIC',
                help='Statistic name to be analysed')
        parser.add_argument(
                'reference', metavar='REFERENCE', type=int,
                help='Reference value for comparison in desired stat unit')
        parser.add_argument(
                '-d', '--timestamp-boundaries',
                metavar=('BEGIN_DATE', 'END_DATE'), nargs=2,
                help='Start and End date in format YYYY:MM:DD hh:mm:ss')
        parser.add_argument(
                '-D', '--start-day', metavar='START_DAY',
                help='Starting time of the day')
        parser.add_argument(
                '-E', '--start-evening', metavar='START_EVENING',
                help='Starting time of the evening')
        parser.add_argument(
                '-N', '--start-night', metavar='START_NIGHT',
                help='starting time of the night')
        parser.add_argument(
                '-l', '--stability-threshold', '--threshold',
                metavar='THRESHOLD', type=int,
                help='Percentage level under which the evolution is considered stable')
        parser.add_argument(
                '-p', '--path-to-file', metavar='PATH_TO_FILE',
                help='Path to XLSX file for evolution calculation')
        parser.add_argument(
                '-u', '--stat-unit',
                metavar='STAT_UNIT', choices=UNIT_OPTION,
                help='Unit of the statistic')
        parser.add_argument(
                '-U', '--table-unit',
                metavar='TABLE_UNIT', choices=UNIT_OPTION,
                help='Unit to show on the table')
        parser.add_argument(
                '-t', '--stat-title',
                help='Statistic names to display on the table')
        parser.add_argument(
                '-w', '--no-suffix', action='store_true',
                help='Do not show statistics with suffixes')
        parser.add_argument(
                '--no_median', action='store_true',
                help='Do not compute median')
        parser.add_argument(
                '--no_mean', action='store_true',
                help='Do not compute mean')

        args = parser.parse_args()
        compute_median = not args.no_median
        compute_mean = not args.no_mean
        stats_with_suffixes = not args.no_suffix

        main(
            args.agent, args.job, args.statistic, args.timestamp_boundaries,
            args.start_day, args.start_evening, args.start_night, args.reference,
            args.stability_threshold, args.stat_unit, args.table_unit,
            args.path_to_file, args.stat_title, compute_median, compute_mean, stats_with_suffixes)
